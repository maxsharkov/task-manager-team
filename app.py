import os
import json
import io
import requests
from flask import Flask, render_template, request, redirect, url_for, jsonify, send_from_directory, send_file
import psycopg2
import psycopg2.extras
from datetime import date, timedelta
import calendar
from openai import OpenAI
from apscheduler.schedulers.background import BackgroundScheduler
from apscheduler.triggers.cron import CronTrigger
import pytz
import gcal
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

app = Flask(__name__)
DATABASE_URL = (
    os.environ.get("DATABASE_URL") or
    os.environ.get("POSTGRES_URL") or
    os.environ.get("POSTGRESQL_URL")
).replace("postgres://", "postgresql://", 1)
openai_client = OpenAI(api_key=os.environ["OPENAI_API_KEY"])

TELEGRAM_TOKEN = os.environ.get("TELEGRAM_BOT_TOKEN")
TELEGRAM_CHAT_ID = os.environ.get("TELEGRAM_CHAT_ID")
WEBHOOK_URL = os.environ.get("WEBHOOK_URL")
NSK = pytz.timezone("Asia/Novosibirsk")

STATUSES = ["Новая", "В работе", "Завершена"]
PRIORITIES = ["Низкий", "Средний", "Высокий"]
ENERGIES = ["Лёгкая", "Средняя", "Тяжёлая"]
CATEGORIES = ["Работа", "Личное"]
RECURRENCES = ["Ежедневно", "Еженедельно", "Ежемесячно"]
STRATEGIC_AREAS = ["Здоровье", "Семья", "Работа", "Прогрессив", "Люди", "Мышление", "Яркость", "Деньги", "Публичность", "Принципы"]

STRATEGIC_SEED = [
    # ── Здоровье ──────────────────────────────────────────────
    ("Проплыть 3 км за 1 час", "Здоровье"),
    ("График спорта: плавание, бег, йога, мышцы — 2-3 раза в неделю", "Здоровье"),
    ("Раз в полгода поясница — УВТ, Регидрон, Регейн, физио", "Здоровье"),
    ("Медитации и дыхание не менее 1 часа в неделю суммарно", "Здоровье"),
    ("Раз в полгода чекап — анализы, обследование", "Здоровье"),
    ("Вести трекер энергии: что заряжает, а что забирает силы", "Здоровье"),
    ("Ввести в рацион: Омега3, Витамин D, B12, клетчатка (овсянка, авокадо, лён, бобовые, овощи)", "Здоровье"),
    ("Регулярный массаж, SPA", "Здоровье"),
    ("Посетить ретрит", "Здоровье"),
    ("Сохранять внутреннюю тишину в перегрузке", "Здоровье"),
    ("Научиться сначала чувствовать тело, потом действовать", "Здоровье"),
    ("Визуализировать себя в 50 лет — семья, улыбка, встречи, загар, учусь, инновации", "Здоровье"),
    ("Научиться проживать момент, а не только достигать цели", "Здоровье"),
    ("Теплый огонь — греть, но не сжигать и не тухнуть", "Здоровье"),
    # ── Семья ─────────────────────────────────────────────────
    ("Раз в неделю семейный вечер без телефонов", "Семья"),
    ("Совместный спорт с детьми — баскетбол, бег, бадминтон, танцы", "Семья"),
    ("Делать сюрпризы близким каждый месяц — подарок, поход, совместная активность", "Семья"),
    ("Раз в полгода большое семейное путешествие", "Семья"),
    ("Раз в год — приключение с родителями", "Семья"),
    ("Учиться вместе с детьми — английский, ИИ, покупки", "Семья"),
    ("Танюшка — совместные сильные эмоции, обучение, приключения", "Семья"),
    ("Разговаривать с детьми как со взрослыми — просить совет, помощь, я-коммуникация", "Семья"),
    ("Присутствовать в моменте с семьей", "Семья"),
    ("Семья — место силы и спокойствия", "Семья"),
    ("Видеть себя в 50 лет счастливым рядом с семьёй", "Семья"),
    # ── Работа ────────────────────────────────────────────────
    ("Настроить серьезный операционный ритм (встречи, темы, цели, обмен опытом)", "Работа"),
    ("Внедрить культуру постконтроля (пообещал и сделал)", "Работа"),
    ("Регулярно вести несколько B2B клиентов (2Gis, ЦФТ, Альфа, Эн+, ИНК)", "Работа"),
    ("Взять проект на РФ (слабости конкурентов, ЦПУ)", "Работа"),
    ("Вести проект вне компании (Альфа, B2B клиент)", "Работа"),
    ("Прожить и доработать локальные стратегии по каждому филиалу — создать систему стратегий", "Работа"),
    ("Доработать процессы — process mining", "Работа"),
    ("Фиксировать примеры про рост бизнеса, антикризис, трансформацию — с людьми/процессами/KPI", "Работа"),
    ("Сделать сложное простым для команды", "Работа"),
    ("Мыслить как CEO, а не функциональный руководитель — фин. устойчивость, акционеры, риски", "Работа"),
    ("Держать в голове главные точки роста бизнеса", "Работа"),
    ("Вдохновлять людей на сверхрезультат", "Работа"),
    ("Каждый день удивлять себя — рывок за пределами того, что могу", "Работа"),
    # ── Технологии ────────────────────────────────────────────
    ("Создать AI-экзоскелет жизни и работы — набор агентов: сказал=сделал, информирование", "Прогрессив"),
    ("Покупать инновационные гаджеты и подписки", "Прогрессив"),
    ("Научиться работать из любой точки мира", "Прогрессив"),
    ("Создать собственный цифровой продукт — сайт, продвижение через Claude, организовать путешествие", "Прогрессив"),
    ("Построить систему управления знаниями — регулярно пересматривать ТОП10 навыков", "Прогрессив"),
    ("Подготовить платформу консультанта/эксперта", "Прогрессив"),
    ("Создать новую карьерную свободу", "Прогрессив"),
    ("Создать следующий уровень возможностей на 10 лет вперёд", "Прогрессив"),
    # ── Люди ──────────────────────────────────────────────────
    ("Каждую неделю общаться с людьми за пределами компании (GMR7) — причём по делу", "Люди"),
    ("Раз в месяц проводить глубокий разговор с сильным человеком", "Люди"),
    ("Составить топ компаний, где хочу оказаться — 2GIS, S7, T2, Росатом, Яндекс, Альфа, Тинькофф, VK, Ozon, Сибур — пообщаться с хантерами", "Люди"),
    ("Растить сильных лидеров и преемников", "Люди"),
    ("Создать сообщество сильных руководителей", "Люди"),
    ("Развивать людей через сложные вызовы", "Люди"),
    ("Формировать вокруг себя сильное окружение", "Люди"),
    ("Научиться быстрее видеть потенциал людей", "Люди"),
    ("Усилить культуру инициативности", "Люди"),
    ("Быть «невидимым коучем» — меньше говорить, больше слушать без намерения", "Люди"),
    ("Стать точкой спокойствия и масштаба для команды", "Люди"),
    ("После общения со мной человек чувствует себя лучше", "Люди"),
    ("Оставлять сильный след через людей", "Люди"),
    # ── Мышление ──────────────────────────────────────────────
    ("Пройти Skolkovo LIFT или аналогичную программу", "Мышление"),
    ("Пройти обучение независимых директоров", "Мышление"),
    ("Общаться с сильными менторами (Торбахов, Пятков, Косолапов, Шоржин, Федорова)", "Мышление"),
    ("Регулярно обновлять ТОП10 навыков будущего", "Мышление"),
    ("Читать сильные книги и исследования — постоянно", "Мышление"),
    ("Найти и развивать менти", "Мышление"),
    ("Ретрит тишины — перезагрузиться", "Мышление"),
    ("Раз в год попадать в совершенно новую среду/культуру (фестиваль, ретрит, поход, музыканты)", "Мышление"),
    # ── Творчество ────────────────────────────────────────────
    ("Путешествие каждые 3 месяца — особенно в новые места", "Яркость"),
    ("Ходить на концерты и выставки", "Яркость"),
    ("Природа — каждую неделю обязательно", "Яркость"),
    ("Раз в месяц мини-путешествие", "Яркость"),
    ("Эндуро раз в месяц", "Яркость"),
    ("Продолжить барабаны и музыку — выступить с песней летом", "Яркость"),
    ("Создать список «Пока не сыграл в ящик» — было и будет", "Яркость"),
    ("Жить часть времени рядом с морем", "Яркость"),
    ("Добавить творчество в повседневность", "Яркость"),
    ("Делать жизнь менее предсказуемой — оставлять место для приключения", "Яркость"),
    # ── Деньги ────────────────────────────────────────────────
    ("Определить финансовую цель на 50 лет — пассивный доход 1 млн руб./мес, состояние $5 млн", "Деньги"),
    ("Раз в полгода перебалансировка портфеля — intelinvest", "Деньги"),
    ("Запустить свой небольшой проект — через Claude Code и интересующихся людей", "Деньги"),
    ("Войти операционно в бизнес — телеком, ИТ, маркетинг, ИИ — знакомые", "Деньги"),
    ("Создать семейное мини-дело", "Деньги"),
    ("Инвестировать в своё имя и влияние — встречи, услуги, менторинг, СМИ", "Деньги"),
    ("Уйти от зависимости только от корпоративной роли", "Деньги"),
    ("Создать свободу жить в любом месте", "Деньги"),
    ("Связать деньги со свободой, а не потреблением", "Деньги"),
    # ── Публичность ───────────────────────────────────────────
    ("Выступать в НГУ и других ВУЗах", "Публичность"),
    ("Выступать на ВЭФ и отраслевых площадках", "Публичность"),
    ("Начать блог или видеоформат — понять, как это делать", "Публичность"),
    ("Каждую неделю выступать публично — театральные курсы пройти", "Публичность"),
    ("Красиво и просто доносить сложные идеи", "Публичность"),
    ("Создать сильный personal brand", "Публичность"),
    ("Создать собственный стиль коммуникации — через отдавать", "Публичность"),
    # ── Система ───────────────────────────────────────────────
    ("Сделать обещания себе священными", "Принципы"),
    ("Разделять, что даёт энергию, а что забирает", "Принципы"),
    ("Каждый день делать то, что боишься — превращать страх в радость", "Принципы"),
    ("Управлять вниманием как главным ресурсом — весь мир внутри меня", "Принципы"),
    ("Отдавать миру — каждый день помогать, поддерживать", "Принципы"),
    ("Осознанность = пауза перед действием", "Принципы"),
    ("Соединить эмоциональное и рациональное «зачем»", "Принципы"),
    ("Быть честным с собой", "Принципы"),
    ("Весь опыт делится на приятный и полезный — думай позитивно", "Принципы"),
    ("Цель — свобода перемещения, общения, занятия", "Принципы"),
    ("Всё проще, чем кажется", "Принципы"),
    ("Амплитуда эмоций", "Принципы"),
    ("Жизнь не такая уж серьёзная штука", "Принципы"),
    ("Я — солнце, светящаяся точка. Свет — отдавать энергию. Точка — всё просто.", "Принципы"),
]


def calc_next_deadline(deadline, recurrence):
    if not deadline or not recurrence:
        return None
    dl = date.fromisoformat(str(deadline))
    if recurrence == "Ежедневно":
        return dl + timedelta(days=1)
    if recurrence == "Еженедельно":
        return dl + timedelta(weeks=1)
    if recurrence == "Ежемесячно":
        month = dl.month % 12 + 1
        year = dl.year + (dl.month // 12)
        day = min(dl.day, calendar.monthrange(year, month)[1])
        return date(year, month, day)
    return None


def get_db():
    return psycopg2.connect(DATABASE_URL)


def init_db():
    with get_db() as conn:
        with conn.cursor() as cur:
            cur.execute("""
                CREATE TABLE IF NOT EXISTS tasks (
                    id SERIAL PRIMARY KEY,
                    title TEXT NOT NULL,
                    status TEXT NOT NULL DEFAULT 'Новая',
                    priority TEXT NOT NULL DEFAULT 'Средний',
                    deadline DATE,
                    created_at DATE DEFAULT CURRENT_DATE,
                    description TEXT,
                    project TEXT,
                    energy TEXT NOT NULL DEFAULT 'Средняя'
                )
            """)
            cur.execute("ALTER TABLE tasks ADD COLUMN IF NOT EXISTS description TEXT")
            cur.execute("ALTER TABLE tasks ADD COLUMN IF NOT EXISTS project TEXT")
            cur.execute("ALTER TABLE tasks ADD COLUMN IF NOT EXISTS energy TEXT NOT NULL DEFAULT 'Средняя'")
            cur.execute("ALTER TABLE tasks ADD COLUMN IF NOT EXISTS tags TEXT")
            cur.execute("ALTER TABLE tasks ADD COLUMN IF NOT EXISTS assignee TEXT")
            cur.execute("ALTER TABLE tasks ADD COLUMN IF NOT EXISTS progress TEXT")
            cur.execute("ALTER TABLE tasks ADD COLUMN IF NOT EXISTS calendar_event_id TEXT")
            cur.execute("ALTER TABLE tasks ADD COLUMN IF NOT EXISTS recurrence TEXT")
            cur.execute("ALTER TABLE tasks ADD COLUMN IF NOT EXISTS closed_at DATE")
            cur.execute("ALTER TABLE tasks ADD COLUMN IF NOT EXISTS category TEXT")
            cur.execute("""
                CREATE TABLE IF NOT EXISTS subtasks (
                    id SERIAL PRIMARY KEY,
                    task_id INTEGER REFERENCES tasks(id) ON DELETE CASCADE,
                    text TEXT NOT NULL,
                    done BOOLEAN DEFAULT FALSE
                )
            """)
            cur.execute("""
                CREATE TABLE IF NOT EXISTS strategic_goals (
                    id SERIAL PRIMARY KEY,
                    title TEXT NOT NULL,
                    area TEXT,
                    created_at DATE DEFAULT CURRENT_DATE
                )
            """)
            cur.execute("""
                CREATE TABLE IF NOT EXISTS strategic_logs (
                    id SERIAL PRIMARY KEY,
                    goal_id INTEGER REFERENCES strategic_goals(id) ON DELETE CASCADE,
                    logged_at DATE DEFAULT CURRENT_DATE,
                    text TEXT NOT NULL
                )
            """)
            cur.execute("""
                CREATE TABLE IF NOT EXISTS strategic_snapshots (
                    id SERIAL PRIMARY KEY,
                    week_date DATE UNIQUE NOT NULL,
                    scores JSONB NOT NULL,
                    review_text TEXT NOT NULL,
                    created_at TIMESTAMP DEFAULT NOW()
                )
            """)
            # Миграция переименований областей
            _RENAMES = [("Технологии", "Прогрессив"), ("Творчество", "Яркость"), ("Система", "Принципы")]
            for old, new in _RENAMES:
                cur.execute("UPDATE strategic_goals SET area=%s WHERE area=%s", (new, old))
                # Переименование ключей в JSONB снимков
                cur.execute("""
                    UPDATE strategic_snapshots
                    SET scores = (scores - %s) || jsonb_build_object(%s, scores->%s)
                    WHERE scores ? %s
                """, (old, new, old, old))
            cur.execute("SELECT COUNT(*) FROM strategic_goals")
            if cur.fetchone()[0] == 0:
                cur.executemany(
                    "INSERT INTO strategic_goals (title, area) VALUES (%s, %s)",
                    STRATEGIC_SEED
                )


def send_telegram(text):
    if not TELEGRAM_TOKEN or not TELEGRAM_CHAT_ID:
        print("Telegram not configured")
        return
    url = f"https://api.telegram.org/bot{TELEGRAM_TOKEN}/sendMessage"
    requests.post(url, json={
        "chat_id": TELEGRAM_CHAT_ID,
        "text": text,
        "parse_mode": "Markdown",
    }, timeout=10)


def build_digest(digest_type="daily"):
    today = date.today()
    today_str = today.isoformat()
    week_ago = (today - timedelta(days=7)).isoformat()

    with get_db() as conn:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            cur.execute("SELECT title, status, priority, deadline, category, closed_at, progress, assignee FROM tasks ORDER BY created_at")
            tasks = cur.fetchall()

    if not tasks:
        send_telegram("📋 *Дайджест задач*\n\nЗадач пока нет.")
        return

    def fmt(t):
        line = f"- [{t['category'] or '?'}] «{t['title']}»"
        if t['deadline']:
            line += f" | дедлайн: {t['deadline']}"
        if t['closed_at']:
            line += f" | закрыта: {t['closed_at']}"
        if t['assignee']:
            line += f" | {t['assignee']}"
        if t['progress']:
            line += f" | прогресс: {t['progress'][:60]}"
        return line

    if digest_type == "weekly":
        closed_period = [t for t in tasks
                         if t['closed_at'] and str(t['closed_at'])[:10] >= week_ago]
        period_label = "за последние 7 дней"
        header = "📊 *Итоги недели — оценка выполнения обещаний*"
        period = "за эту неделю"
    else:
        closed_period = [t for t in tasks
                         if t['closed_at'] and str(t['closed_at'])[:10] == today_str]
        period_label = "за сегодня"
        header = "🌙 *Дайджест — оценка выполнения обещаний*"
        period = "за сегодня"

    overdue = [t for t in tasks if t['status'] != 'Завершена' and t['deadline']
               and str(t['deadline']) < today_str]
    active_hp = [t for t in tasks if t['status'] != 'Завершена' and t['priority'] == 'Высокий'
                 and (not t['deadline'] or str(t['deadline']) >= today_str)]

    sections = []
    sections.append(
        f"✅ Завершено {period_label}:\n" + "\n".join(fmt(t) for t in closed_period)
        if closed_period else f"✅ Завершено {period_label}: ничего"
    )
    sections.append(
        "🔴 Просрочено (активные):\n" + "\n".join(fmt(t) for t in overdue)
        if overdue else "🔴 Просрочено: нет"
    )
    sections.append(
        "🔥 В работе (высокий приоритет):\n" + "\n".join(fmt(t) for t in active_hp)
        if active_hp else "🔥 В работе (высокий приоритет): нет"
    )

    tasks_structured = "\n\n".join(sections)

    prompt = f"""Ты жёсткий, честный коуч топ-менеджера. Без корпоративного языка, без комплиментов. Только факты и прямая оценка.

Сегодня {today_str}. Задачи структурированы по факту:

{tasks_structured}

ВАЖНО: раздел «Завершено» — это реально выполненные задачи. Учитывай их при выставлении оценки.

Сделай оценку выполнения обещаний {period} по двум категориям.

🏢 РАБОЧИЕ ЦЕЛИ (категория Работа)
1. Что обещано / в работе
2. Что реально выполнено (из раздела Завершено)
3. Что просрочено или зависло
4. Оценка: X/10

👤 ЛИЧНЫЕ ЦЕЛИ (категория Личное)
1. Что обещано / в работе
2. Что реально выполнено (из раздела Завершено)
3. Что просрочено или зависло
4. Оценка: X/10

📌 ИТОГО
— Общая оценка: X/10
— Главный паттерн: что системно не выполняется
— Одна конкретная рекомендация

Шкала: 9-10 = отлично, 7-8 = хорошо, 5-6 = средне, ниже 5 = провал. Будь честным."""

    response = openai_client.chat.completions.create(
        model="gpt-4o",
        messages=[
            {
                "role": "system",
                "content": "Ты жёсткий честный коуч. Отвечай на русском, структурированно, без воды и лести."
            },
            {
                "role": "user",
                "content": prompt
            }
        ]
    )

    summary = response.choices[0].message.content
    send_telegram(f"{header}\n\n{summary}")


def daily_digest():
    build_digest("daily")


def weekly_digest():
    build_digest("weekly")


def build_strategic_digest():
    today = date.today()
    today_str = today.isoformat()

    _AREA_EMOJI = {
        "Здоровье": "💪", "Семья": "❤️", "Работа": "⚡", "Прогрессив": "🤖",
        "Люди": "🤝", "Мышление": "🧠", "Яркость": "🎨", "Деньги": "💰",
        "Публичность": "📣", "Принципы": "⚙️",
    }

    with get_db() as conn:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            cur.execute("SELECT * FROM strategic_goals ORDER BY area, title")
            goals = cur.fetchall()
            if not goals:
                send_telegram("🧭 *Стратегический обзор*\n\nСтратегических целей пока нет.")
                return
            goal_logs = {}
            for g in goals:
                cur.execute("""
                    SELECT text, logged_at FROM strategic_logs
                    WHERE goal_id = %s ORDER BY logged_at DESC, id DESC LIMIT 2
                """, (g['id'],))
                goal_logs[g['id']] = cur.fetchall()

    # Группировка по областям
    from collections import defaultdict, OrderedDict
    areas = OrderedDict()
    for area in STRATEGIC_AREAS:
        areas[area] = []
    for g in goals:
        area = g['area'] or "Прочее"
        if area not in areas:
            areas[area] = []
        areas[area].append(g)

    block_lines = []
    for area, area_goals in areas.items():
        if not area_goals:
            continue
        active_7 = []    # цели с записью за последние 7 дней
        active_30 = []   # цели с записью за 8–30 дней
        silent = []      # цели без единой записи или молчат 30+ дней

        for g in area_goals:
            logs = goal_logs.get(g['id'], [])
            if logs:
                last_date = logs[0]['logged_at']
                days_ago = (today - last_date).days
                if days_ago <= 7:
                    active_7.append((g, logs, days_ago))
                elif days_ago <= 30:
                    active_30.append((g, logs, days_ago))
                else:
                    silent.append(g)
            else:
                silent.append(g)

        emoji = _AREA_EMOJI.get(area, "•")
        lines = [f"БЛОК: {emoji} {area} ({len(area_goals)} целей)"]
        lines.append(f"Активных за 7 дней: {len(active_7)}")
        for g, logs, days_ago in active_7:
            for l in logs[:2]:
                lines.append(f"  - «{g['title']}» [{l['logged_at']}]: {l['text'][:80]}")
        lines.append(f"Молчат 8-30 дней: {len(active_30)}")
        if active_30:
            lines.append("  " + ", ".join(f"«{g['title']}»" for g, _, _ in active_30[:3]))
        lines.append(f"Застой / нет записей: {len(silent)}")
        block_lines.append("\n".join(lines))

    goals_structured = "\n\n".join(block_lines)

    prompt = f"""Ты честный жёсткий коуч топ-менеджера. Воскресный стратегический обзор.

Сегодня {today_str}.

Данные по 10 стратегическим блокам:

{goals_structured}

Задача: оцени каждый блок по шкале 0–10, где:
- 0-3 → полный застой (нет активности)
- 4-6 → слабое движение (1-2 цели из блока)
- 7-8 → нормальный темп (половина целей активна)
- 9-10 → сильное движение (большинство целей с записями за неделю)

Структура ответа — строго в таком формате:

```json
{{"Здоровье": 0, "Семья": 0, "Работа": 0, "Прогрессив": 0, "Люди": 0, "Мышление": 0, "Яркость": 0, "Деньги": 0, "Публичность": 0, "Принципы": 0}}
```

🧭 СТРАТЕГИЧЕСКИЙ ОБЗОР — {today_str}

[для каждого блока]:
{{эмодзи}} {{Название}} — {{X}}/10 {{стрелка ↑/→/↓}} — {{одна фраза: оценка блока}}
[если есть активные записи за 7 дней — перечисли только тексты записей, без названий целей, не более 4 на блок]:
  • {{текст записи}} [{{дата}}]

Если активных записей нет — строку с буллетами не добавляй.
Эмодзи для блоков: Здоровье💪 Семья❤️ Работа⚡ Прогрессив🤖 Люди🤝 Мышление🧠 Яркость🎨 Деньги💰 Публичность📣 Принципы⚙️

📊 ИТОГ НЕДЕЛИ — средняя оценка по всем блокам, 1 предложение про общий тренд.

📌 ГЛАВНЫЙ ВОПРОС — один сильный вопрос по самому застывшему блоку.

Без воды. Без корпоративного языка. Максимально конкретно."""

    response = openai_client.chat.completions.create(
        model="gpt-4o",
        messages=[
            {"role": "system", "content": "Ты жёсткий честный коуч. Отвечай на русском, без воды."},
            {"role": "user", "content": prompt}
        ]
    )
    raw = response.choices[0].message.content

    # Извлекаем JSON с оценками из блока ```json ... ```
    import re as _re
    import json as _json
    scores = {}
    json_match = _re.search(r'```json\s*(\{.*?\})\s*```', raw, _re.DOTALL)
    if json_match:
        try:
            scores = _json.loads(json_match.group(1))
        except Exception:
            pass
    # Нарратив — всё после блока с JSON
    narrative = _re.sub(r'```json\s*\{.*?\}\s*```\s*', '', raw, flags=_re.DOTALL).strip()

    # Сохраняем снимок в БД (upsert по week_date)
    if scores:
        with get_db() as conn:
            with conn.cursor() as cur:
                cur.execute("""
                    INSERT INTO strategic_snapshots (week_date, scores, review_text)
                    VALUES (%s, %s, %s)
                    ON CONFLICT (week_date) DO UPDATE
                        SET scores = EXCLUDED.scores,
                            review_text = EXCLUDED.review_text,
                            created_at = NOW()
                """, (today, _json.dumps(scores, ensure_ascii=False), narrative))

    send_telegram(f"🧭 *Стратегический обзор*\n\n{narrative}")


def strategic_review():
    build_strategic_digest()


# ── Планировщик ────────────────────────────────────────────
scheduler = BackgroundScheduler(timezone=NSK)
scheduler.add_job(daily_digest, CronTrigger(hour=21, minute=45, timezone=NSK))
scheduler.add_job(weekly_digest, CronTrigger(day_of_week="sun", hour=20, minute=0, timezone=NSK))
scheduler.add_job(strategic_review, CronTrigger(day_of_week="sun", hour=10, minute=0, timezone=NSK))
scheduler.start()


def parse_tags(raw):
    if not raw:
        return []
    return [t.strip().lstrip("#") for t in raw.replace(",", " ").split() if t.strip()]


def normalize_tags(raw):
    tags = parse_tags(raw)
    return ",".join(tags) if tags else None


@app.route("/tags")
def get_tags():
    with get_db() as conn:
        with conn.cursor() as cur:
            cur.execute("SELECT tags FROM tasks WHERE tags IS NOT NULL AND tags != ''")
            rows = cur.fetchall()
    all_tags = set()
    for row in rows:
        all_tags.update(parse_tags(row[0]))
    return jsonify(sorted(all_tags))


@app.route("/")
def index():
    priority_filter = request.args.get("priority", "")
    search_query = request.args.get("q", "").strip()

    query = """
        SELECT t.*,
            COUNT(s.id) AS subtask_total,
            COUNT(s.id) FILTER (WHERE s.done) AS subtask_done
        FROM tasks t
        LEFT JOIN subtasks s ON s.task_id = t.id
        WHERE t.status != 'Завершена'
    """
    params = []

    if priority_filter:
        query += " AND t.priority = %s"
        params.append(priority_filter)
    if search_query:
        query += " AND (t.title ILIKE %s OR t.assignee ILIKE %s)"
        params.extend([f"%{search_query}%"] * 2)

    query += """
        GROUP BY t.id
        ORDER BY
            t.category NULLS LAST,
            CASE t.priority WHEN 'Высокий' THEN 1 WHEN 'Средний' THEN 2 ELSE 3 END,
            t.deadline ASC NULLS LAST
    """

    with get_db() as conn:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            cur.execute(query, params)
            tasks = cur.fetchall()

            cur.execute("""
                SELECT * FROM tasks WHERE status = 'Завершена'
                ORDER BY closed_at DESC NULLS LAST, id DESC
            """)
            done_tasks = cur.fetchall()

            cur.execute("""
                SELECT g.*,
                    MAX(l.logged_at) AS last_log_date
                FROM strategic_goals g
                LEFT JOIN strategic_logs l ON l.goal_id = g.id
                GROUP BY g.id
                ORDER BY g.area, g.title
            """)
            strategic_goals = cur.fetchall()

            # Загружаем все логи по каждой цели
            goal_logs = {}
            for g in strategic_goals:
                cur.execute("""
                    SELECT text, logged_at FROM strategic_logs
                    WHERE goal_id = %s ORDER BY logged_at DESC, id DESC
                """, (g['id'],))
                goal_logs[g['id']] = cur.fetchall()

    return render_template(
        "index.html",
        tasks=tasks,
        done_tasks=done_tasks,
        strategic_goals=strategic_goals,
        goal_logs=goal_logs,
        statuses=STATUSES,
        priorities=PRIORITIES,
        energies=ENERGIES,
        recurrences=RECURRENCES,
        categories=CATEGORIES,
        strategic_areas=STRATEGIC_AREAS,
        priority_filter=priority_filter,
        search_query=search_query,
        today=date.today(),
        parse_tags=parse_tags,
    )


@app.route("/strategy/add", methods=["POST"])
def strategy_add():
    title = request.form.get("title", "").strip()
    area = request.form.get("area", "")
    if title:
        with get_db() as conn:
            with conn.cursor() as cur:
                cur.execute(
                    "INSERT INTO strategic_goals (title, area) VALUES (%s, %s)",
                    (title, area)
                )
    return redirect("/?tab=strategy")


@app.route("/strategy/log/<int:goal_id>", methods=["POST"])
def strategy_log(goal_id):
    text = request.form.get("text", "").strip()
    if text:
        with get_db() as conn:
            with conn.cursor() as cur:
                cur.execute(
                    "INSERT INTO strategic_logs (goal_id, text) VALUES (%s, %s)",
                    (goal_id, text)
                )
    return redirect("/?tab=strategy")


@app.route("/strategy/edit/<int:goal_id>", methods=["POST"])
def strategy_edit(goal_id):
    title = request.form.get("title", "").strip()
    area = request.form.get("area", "")
    if title:
        with get_db() as conn:
            with conn.cursor() as cur:
                cur.execute(
                    "UPDATE strategic_goals SET title=%s, area=%s WHERE id=%s",
                    (title, area, goal_id)
                )
    return redirect("/?tab=strategy")


@app.route("/strategy/delete/<int:goal_id>", methods=["POST"])
def strategy_delete(goal_id):
    with get_db() as conn:
        with conn.cursor() as cur:
            cur.execute("DELETE FROM strategic_goals WHERE id=%s", (goal_id,))
    return redirect("/?tab=strategy")


@app.route("/gcal-resync", methods=["POST"])
def gcal_resync():
    """Пересоздаёт Calendar-события для активных повторяющихся задач с RRULE.
    Если дедлайн в прошлом — сдвигает до ближайшего предстоящего."""
    updated = 0
    errors = []
    today_str = date.today().isoformat()
    with get_db() as conn:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            cur.execute("""
                SELECT id, title, deadline, priority, assignee, recurrence, calendar_event_id
                FROM tasks
                WHERE status != 'Завершена' AND recurrence IS NOT NULL AND deadline IS NOT NULL
            """)
            tasks = cur.fetchall()
            for t in tasks:
                # Удаляем старое событие
                if t["calendar_event_id"]:
                    gcal.delete_event(t["calendar_event_id"])
                # Если дедлайн в прошлом — двигаем до ближайшего будущего
                dl = t["deadline"]
                while dl and str(dl) < today_str:
                    dl = calc_next_deadline(dl, t["recurrence"])
                if not dl:
                    continue
                new_id = gcal.create_event(
                    t["title"], dl, t["priority"],
                    None, t["assignee"], t["recurrence"],
                    date.today()   # события с сегодня, UNTIL=дедлайн
                )
                cur.execute(
                    "UPDATE tasks SET calendar_event_id=%s, deadline=%s WHERE id=%s",
                    (new_id, dl, t["id"])
                )
                if new_id:
                    updated += 1
                else:
                    errors.append(t["title"])
    return jsonify({"ok": True, "updated": updated, "errors": errors})


@app.route("/gcal-cleanup", methods=["POST"])
def gcal_cleanup():
    """Удаляет Calendar-события у всех завершённых задач."""
    cleaned = 0
    with get_db() as conn:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            cur.execute("""
                SELECT id, calendar_event_id FROM tasks
                WHERE status = 'Завершена' AND calendar_event_id IS NOT NULL
            """)
            tasks = cur.fetchall()
            for t in tasks:
                gcal.delete_event(t["calendar_event_id"])
                cur.execute("UPDATE tasks SET calendar_event_id=NULL WHERE id=%s", (t["id"],))
                cleaned += 1
    return jsonify({"ok": True, "cleaned": cleaned})


@app.route("/strategy/digest", methods=["POST"])
def strategy_digest_manual():
    try:
        build_strategic_digest()
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


@app.route("/strategy/analyze", methods=["POST"])
def strategy_analyze():
    today = date.today().isoformat()
    with get_db() as conn:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            cur.execute("SELECT * FROM strategic_goals ORDER BY area, title")
            goals = cur.fetchall()
            if not goals:
                return jsonify({"summary": "Стратегических целей пока нет."})
            goal_logs = {}
            for g in goals:
                cur.execute("""
                    SELECT text, logged_at FROM strategic_logs
                    WHERE goal_id = %s ORDER BY logged_at ASC, id ASC
                """, (g['id'],))
                goal_logs[g['id']] = cur.fetchall()

    lines = []
    for g in goals:
        logs = goal_logs.get(g['id'], [])
        if logs:
            log_entries = "\n".join(f"  [{l['logged_at']}] {l['text']}" for l in logs)
        else:
            log_entries = "  Записей нет"
        lines.append(f"[{g['area']}] «{g['title']}»\n{log_entries}")

    goals_text = "\n\n".join(lines)

    response = openai_client.chat.completions.create(
        model="gpt-4o",
        messages=[
            {
                "role": "system",
                "content": "Ты жёсткий честный коуч топ-менеджера. Отвечай на русском, структурированно, без воды и лести."
            },
            {
                "role": "user",
                "content": (
                    f"Сегодня {today}. Стратегические цели и история всех записей:\n\n{goals_text}\n\n"
                    "Дай полный анализ:\n"
                    "1. По каждой области (Бизнес, Карьера, Люди, Нетворкинг, Личное) — краткий статус: что движется, что стоит.\n"
                    "2. Топ-3 цели с наилучшим прогрессом — отметь конкретно.\n"
                    "3. Топ-3 застывших цели — назови следующее конкретное действие для каждой.\n"
                    "4. Главный паттерн: что системно игнорируется.\n"
                    "5. Одна рекомендация на эту неделю — конкретная и измеримая."
                )
            }
        ]
    )

    return jsonify({"summary": response.choices[0].message.content})


def _goal_activity_status(last_log_date, today):
    if not last_log_date:
        return "Нет записей"
    days_ago = (today - last_log_date).days
    if days_ago <= 7:
        return "Активна"
    if days_ago <= 14:
        return "Замедление"
    return "Стагнация"


@app.route("/strategy/export")
def strategy_export():
    today = date.today()
    with get_db() as conn:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            cur.execute("""
                SELECT g.*, MAX(l.logged_at) AS last_log_date, COUNT(l.id) AS log_count
                FROM strategic_goals g
                LEFT JOIN strategic_logs l ON l.goal_id = g.id
                GROUP BY g.id
                ORDER BY g.area, g.title
            """)
            goals = cur.fetchall()

            cur.execute("""
                SELECT l.goal_id, l.logged_at, l.text, g.area, g.title
                FROM strategic_logs l
                JOIN strategic_goals g ON g.id = l.goal_id
                ORDER BY g.area, g.title, l.logged_at DESC, l.id DESC
            """)
            logs = cur.fetchall()

    wb = Workbook()

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1D1D1F", end_color="1D1D1F", fill_type="solid")
    wrap = Alignment(wrap_text=True, vertical="top")

    def style_header(ws, ncols):
        for col in range(1, ncols + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(vertical="center")
        ws.freeze_panes = "A2"

    ws1 = wb.active
    ws1.title = "Цели"
    ws1.append(["Область", "Цель", "Статус", "Дней с последней записи", "Последняя запись", "Всего записей", "Создана"])
    for g in goals:
        days_ago = (today - g["last_log_date"]).days if g["last_log_date"] else None
        ws1.append([
            g["area"],
            g["title"],
            _goal_activity_status(g["last_log_date"], today),
            days_ago if days_ago is not None else "",
            g["last_log_date"].isoformat() if g["last_log_date"] else "",
            g["log_count"],
            g["created_at"].isoformat() if g["created_at"] else "",
        ])
    widths1 = [14, 55, 14, 20, 16, 14, 12]
    for i, w in enumerate(widths1, start=1):
        ws1.column_dimensions[get_column_letter(i)].width = w
    style_header(ws1, len(widths1))

    ws2 = wb.create_sheet("История записей")
    ws2.append(["Область", "Цель", "Дата", "Комментарий"])
    for l in logs:
        ws2.append([l["area"], l["title"], l["logged_at"].isoformat(), l["text"]])
    for row in ws2.iter_rows(min_row=2):
        row[3].alignment = wrap
    widths2 = [14, 45, 12, 90]
    for i, w in enumerate(widths2, start=1):
        ws2.column_dimensions[get_column_letter(i)].width = w
    style_header(ws2, len(widths2))

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"strategy_export_{today.isoformat()}.xlsx"
    return send_file(
        buf,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.route("/add", methods=["POST"])
def add():
    title = request.form.get("title", "").strip()
    if not title:
        return redirect(url_for("index"))

    status = request.form.get("status", "Новая")
    priority = request.form.get("priority", "Средний")
    deadline = request.form.get("deadline") or None
    project = request.form.get("project", "").strip() or None
    energy = request.form.get("energy", "Средняя")
    assignee = request.form.get("assignee", "").strip() or None
    recurrence = request.form.get("recurrence", "").strip() or None
    category = request.form.get("category", "Работа")

    gcal_start = date.today() if recurrence else None
    event_id = gcal.create_event(title, deadline, priority, None, assignee, recurrence, gcal_start) if deadline else None

    with get_db() as conn:
        with conn.cursor() as cur:
            cur.execute(
                "INSERT INTO tasks (title, status, priority, deadline, energy, assignee, calendar_event_id, recurrence, category)"
                " VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)",
                (title, status, priority, deadline, energy, assignee, event_id, recurrence, category),
            )
    return redirect(url_for("index"))


@app.route("/update/<int:task_id>", methods=["POST"])
def update(task_id):
    status = request.form.get("status")
    with get_db() as conn:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            cur.execute("SELECT * FROM tasks WHERE id=%s", (task_id,))
            task = cur.fetchone()
            if not task:
                return redirect(url_for("index"))

            was_completed = task["status"] == "Завершена"
            now_completed = status == "Завершена"
            closed_at = date.today() if now_completed else None

            cur.execute("UPDATE tasks SET status=%s, closed_at=%s WHERE id=%s",
                        (status, closed_at, task_id))

            if now_completed and not was_completed:
                # Удаляем Calendar-событие завершённой задачи
                if task["calendar_event_id"]:
                    gcal.delete_event(task["calendar_event_id"])
                    cur.execute("UPDATE tasks SET calendar_event_id=NULL WHERE id=%s", (task_id,))

                # Для повторяющихся — создаём следующий инстанс
                if task["recurrence"]:
                    next_dl = calc_next_deadline(task["deadline"], task["recurrence"])
                    new_event_id = gcal.create_event(
                        task["title"], next_dl, task["priority"],
                        None, task["assignee"], task["recurrence"],
                        date.today()   # события начинаются с сегодня, не с дедлайна
                    ) if next_dl else None
                    cur.execute(
                        "INSERT INTO tasks (title, status, priority, deadline, energy,"
                        " assignee, recurrence, calendar_event_id, category)"
                        " VALUES (%s,'Новая',%s,%s,%s,%s,%s,%s,%s)",
                        (task["title"], task["priority"], next_dl,
                         task["energy"], task["assignee"], task["recurrence"],
                         new_event_id, task["category"])
                    )
    return redirect(url_for("index"))


@app.route("/edit/<int:task_id>", methods=["GET"])
def edit(task_id):
    with get_db() as conn:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            cur.execute("SELECT * FROM tasks WHERE id = %s", (task_id,))
            task = cur.fetchone()
    if not task:
        return redirect(url_for("index"))
    with get_db() as conn:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            cur.execute("SELECT * FROM subtasks WHERE task_id=%s ORDER BY id", (task_id,))
            subtasks = cur.fetchall()
    return render_template("edit.html", task=task, subtasks=subtasks,
                           statuses=STATUSES, priorities=PRIORITIES, energies=ENERGIES,
                           recurrences=RECURRENCES, categories=CATEGORIES,
                           today=date.today(), parse_tags=parse_tags)


@app.route("/edit/<int:task_id>", methods=["POST"])
def edit_save(task_id):
    title = request.form.get("title", "").strip()
    if not title:
        return redirect(url_for("edit", task_id=task_id))

    status = request.form.get("status", "Новая")
    priority = request.form.get("priority", "Средний")
    deadline = request.form.get("deadline") or None
    energy = request.form.get("energy", "Средняя")
    assignee = request.form.get("assignee", "").strip() or None
    progress = request.form.get("progress", "").strip() or None
    recurrence = request.form.get("recurrence", "").strip() or None
    category = request.form.get("category", "Работа")

    with get_db() as conn:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            cur.execute("SELECT calendar_event_id, status, closed_at FROM tasks WHERE id=%s", (task_id,))
            row = cur.fetchone()

    existing_event_id = row["calendar_event_id"] if row else None
    new_event_id = existing_event_id

    gcal_start = date.today() if recurrence else None
    if deadline:
        if existing_event_id:
            gcal.update_event(existing_event_id, title, deadline, priority, None, assignee, recurrence, gcal_start)
        else:
            new_event_id = gcal.create_event(title, deadline, priority, None, assignee, recurrence, gcal_start)
    else:
        if existing_event_id:
            gcal.delete_event(existing_event_id)
            new_event_id = None

    closed_at = None
    if row:
        if status == "Завершена" and row["status"] != "Завершена":
            closed_at = date.today()
            # Удаляем Calendar-событие при завершении через редактирование
            if new_event_id:
                gcal.delete_event(new_event_id)
                new_event_id = None
        elif status == "Завершена":
            closed_at = row["closed_at"]

    with get_db() as conn:
        with conn.cursor() as cur:
            cur.execute(
                "UPDATE tasks SET title=%s, status=%s, priority=%s, deadline=%s,"
                " energy=%s, assignee=%s, progress=%s, calendar_event_id=%s, recurrence=%s,"
                " closed_at=%s, category=%s WHERE id=%s",
                (title, status, priority, deadline, energy, assignee, progress,
                 new_event_id, recurrence, closed_at, category, task_id),
            )
    return redirect(url_for("index"))


@app.route("/delete/<int:task_id>", methods=["POST"])
def delete(task_id):
    with get_db() as conn:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            cur.execute("SELECT calendar_event_id FROM tasks WHERE id=%s", (task_id,))
            row = cur.fetchone()
            if row and row["calendar_event_id"]:
                gcal.delete_event(row["calendar_event_id"])
            cur.execute("DELETE FROM tasks WHERE id = %s", (task_id,))
    return redirect(url_for("index"))


@app.route("/voice", methods=["POST"])
def voice():
    audio_file = request.files.get("audio")
    if not audio_file:
        return jsonify({"error": "Нет аудио"}), 400

    transcript = openai_client.audio.transcriptions.create(
        model="whisper-1",
        file=("audio.webm", audio_file.stream, audio_file.mimetype or "audio/webm"),
        language="ru",
    )
    text = transcript.text

    today = date.today().isoformat()
    response = openai_client.chat.completions.create(
        model="gpt-4o",
        messages=[
            {
                "role": "system",
                "content": (
                    "Ты помощник, который извлекает поля задачи из голосового сообщения. "
                    "Верни ТОЛЬКО валидный JSON без markdown-блоков. "
                    "Поля: title (строка), priority (одно из: Низкий, Средний, Высокий), "
                    "status (одно из: Новая, В работе, Завершена), "
                    "deadline (дата YYYY-MM-DD или null), "
                    "description (краткое описание или контекст задачи, строка или null), "
                    "project (название проекта или направления, строка или null), "
                    "energy (одно из: Лёгкая, Средняя, Тяжёлая — оцени по сложности задачи), "
                    "assignee (кто должен сделать задачу, строка или null), "
                    "category (одно из: Работа, Личное — определи по смыслу задачи). "
                    "Если приоритет не упомянут — Средний. Если статус не упомянут — Новая. "
                    "Если энергия не упомянута — оцени самостоятельно по смыслу задачи."
                )
            },
            {
                "role": "user",
                "content": f"Сегодня {today}. Голосовое сообщение: «{text}»"
            }
        ],
        response_format={"type": "json_object"},
    )

    fields = json.loads(response.choices[0].message.content)
    fields["transcript"] = text
    return jsonify(fields)


def tg_reply(chat_id, text):
    requests.post(
        f"https://api.telegram.org/bot{TELEGRAM_TOKEN}/sendMessage",
        json={"chat_id": chat_id, "text": text, "parse_mode": "Markdown"},
        timeout=10,
    )


def get_tg_file(file_id):
    info = requests.get(
        f"https://api.telegram.org/bot{TELEGRAM_TOKEN}/getFile",
        params={"file_id": file_id}, timeout=10
    ).json()
    file_path = info["result"]["file_path"]
    audio = requests.get(
        f"https://api.telegram.org/file/bot{TELEGRAM_TOKEN}/{file_path}", timeout=30
    )
    return audio.content


def get_all_tasks_text():
    with get_db() as conn:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            cur.execute("SELECT id, title, status, priority, deadline, project, tags, energy FROM tasks ORDER BY created_at")
            tasks = cur.fetchall()
    if not tasks:
        return None, []
    lines = []
    for t in tasks:
        line = f"[{t['id']}] «{t['title']}» | {t['status']} | {t['priority']}"
        if t['project']:
            line += f" | {t['project']}"
        if t['deadline']:
            line += f" | до {t['deadline']}"
        if t['tags']:
            line += f" | #{' #'.join(parse_tags(t['tags']))}"
        lines.append(line)
    return "\n".join(lines), tasks


def process_bot_message(text):
    today = date.today().isoformat()
    tasks_text, _ = get_all_tasks_text()
    tasks_context = f"Список задач:\n{tasks_text}" if tasks_text else "Задач пока нет."

    response = openai_client.chat.completions.create(
        model="gpt-4o",
        messages=[
            {
                "role": "system",
                "content": (
                    f"Ты умный ассистент таск-менеджера. Сегодня {today}.\n"
                    f"{tasks_context}\n\n"
                    "Ты можешь:\n"
                    "1. Отвечать на вопросы о задачах\n"
                    "2. Добавлять новые задачи\n"
                    "3. Давать дайджест/статус\n\n"
                    "Верни JSON: {\"action\": \"answer\"|\"add_task\", \"text\": \"...\", "
                    "\"task\": {\"title\": ..., \"priority\": \"Низкий|Средний|Высокий\", "
                    "\"status\": \"Новая|В работе|Завершена\", \"deadline\": \"YYYY-MM-DD или null\", "
                    "\"project\": \"...\", \"assignee\": \"...\", \"energy\": \"Лёгкая|Средняя|Тяжёлая\"}}.\n"
                    "action=add_task если пользователь хочет добавить задачу. "
                    "action=answer для всего остального. "
                    "text — ответ пользователю на русском, кратко."
                )
            },
            {"role": "user", "content": text}
        ],
        response_format={"type": "json_object"},
    )

    result = json.loads(response.choices[0].message.content)

    if result.get("action") == "add_task" and result.get("task"):
        t = result["task"]
        title = t.get("title", "Без названия")
        deadline = t.get("deadline") or None
        priority = t.get("priority", "Средний")
        assignee = t.get("assignee") or None
        event_id = gcal.create_event(title, deadline, priority, None, assignee) if deadline else None
        with get_db() as conn:
            with conn.cursor() as cur:
                cur.execute(
                    "INSERT INTO tasks (title, status, priority, deadline, energy, assignee, calendar_event_id)"
                    " VALUES (%s,%s,%s,%s,%s,%s,%s)",
                    (
                        title,
                        t.get("status", "Новая"),
                        priority,
                        deadline,
                        t.get("energy", "Средняя"),
                        assignee,
                        event_id,
                    )
                )
        return result.get("text", "✅ Задача добавлена")

    return result.get("text", "Не понял запрос — попробуй ещё раз")


@app.route("/webhook", methods=["POST"])
def webhook():
    if not TELEGRAM_TOKEN:
        return jsonify({"ok": False}), 400

    update = request.json or {}
    message = update.get("message", {})
    chat_id = message.get("chat", {}).get("id")
    if not chat_id:
        return jsonify({"ok": True})

    # Только от владельца
    if str(chat_id) != str(TELEGRAM_CHAT_ID):
        tg_reply(chat_id, "Нет доступа.")
        return jsonify({"ok": True})

    try:
        # Голосовое сообщение
        if "voice" in message:
            tg_reply(chat_id, "🎙 Обрабатываю...")
            audio_bytes = get_tg_file(message["voice"]["file_id"])
            import io
            transcript = openai_client.audio.transcriptions.create(
                model="whisper-1",
                file=("voice.ogg", io.BytesIO(audio_bytes), "audio/ogg"),
                language="ru",
            )
            text = transcript.text
            reply = process_bot_message(text)
            tg_reply(chat_id, f"🎙 «{text}»\n\n{reply}")

        # Текстовое сообщение
        elif "text" in message:
            text = message["text"]
            if text == "/start":
                tg_reply(chat_id, "👋 Привет! Я таск-менеджер.\n\nМогу:\n• Показать задачи\n• Добавить задачу голосом или текстом\n• Дать дайджест\n\nПросто напиши или надиктуй.")
            else:
                reply = process_bot_message(text)
                tg_reply(chat_id, reply)

    except Exception as e:
        tg_reply(chat_id, f"⚠️ Ошибка: {str(e)[:200]}")

    return jsonify({"ok": True})


@app.route("/setup-webhook")
def setup_webhook():
    if not TELEGRAM_TOKEN or not WEBHOOK_URL:
        return jsonify({"error": "TELEGRAM_BOT_TOKEN или WEBHOOK_URL не настроены"}), 400
    res = requests.post(
        f"https://api.telegram.org/bot{TELEGRAM_TOKEN}/setWebhook",
        json={"url": WEBHOOK_URL},
        timeout=10,
    ).json()
    return jsonify(res)


@app.route("/ai-search", methods=["POST"])
def ai_search():
    query = (request.json or {}).get("query", "").strip()
    if not query:
        return jsonify({"ids": [], "explanation": "Пустой запрос"}), 400

    with get_db() as conn:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            cur.execute("SELECT id, title, status, priority, deadline, description, project, tags FROM tasks ORDER BY created_at")
            tasks = cur.fetchall()

    if not tasks:
        return jsonify({"ids": [], "explanation": "Задач нет"})

    tasks_text = "\n".join([
        f"[{t['id']}] «{t['title']}»"
        + (f" | проект: {t['project']}" if t['project'] else "")
        + (f" | теги: {t['tags']}" if t['tags'] else "")
        + (f" | {t['description'][:100]}" if t['description'] else "")
        + f" | {t['status']} | {t['priority']}"
        for t in tasks
    ])

    response = openai_client.chat.completions.create(
        model="gpt-4o",
        messages=[
            {
                "role": "system",
                "content": (
                    "Ты помощник для поиска задач. Найди задачи, которые соответствуют запросу по смыслу. "
                    "Верни JSON: {\"ids\": [список id], \"explanation\": \"краткое объяснение на русском\"}. "
                    "ids — массив числовых id подходящих задач. Если ничего не найдено — пустой массив."
                )
            },
            {
                "role": "user",
                "content": f"Запрос: «{query}»\n\nЗадачи:\n{tasks_text}"
            }
        ],
        response_format={"type": "json_object"},
    )

    result = json.loads(response.choices[0].message.content)
    return jsonify(result)


@app.route("/api/tactic-stats")
def tactic_stats():
    today = date.today()
    week_ago  = today - timedelta(days=7)
    month_ago = today - timedelta(days=30)
    with get_db() as conn:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            cur.execute("SELECT status, COUNT(*) AS cnt FROM tasks GROUP BY status")
            by_status = {r["status"]: r["cnt"] for r in cur.fetchall()}
            cur.execute("SELECT priority, COUNT(*) AS cnt FROM tasks WHERE status != 'Завершена' GROUP BY priority")
            by_priority = {r["priority"]: r["cnt"] for r in cur.fetchall()}
            cur.execute("SELECT COUNT(*) AS cnt FROM tasks WHERE status='Завершена' AND closed_at >= %s", (week_ago,))
            closed_7 = cur.fetchone()["cnt"]
            cur.execute("SELECT COUNT(*) AS cnt FROM tasks WHERE status='Завершена' AND closed_at >= %s", (month_ago,))
            closed_30 = cur.fetchone()["cnt"]
            cur.execute("SELECT COUNT(*) AS cnt FROM tasks WHERE deadline < %s AND status != 'Завершена'", (today,))
            overdue = cur.fetchone()["cnt"]
            cur.execute("""
                SELECT project, COUNT(*) AS cnt FROM tasks
                WHERE status != 'Завершена' AND project IS NOT NULL AND project != ''
                GROUP BY project ORDER BY cnt DESC LIMIT 8
            """)
            by_project = [dict(r) for r in cur.fetchall()]
            cur.execute("SELECT COUNT(*) AS cnt FROM tasks WHERE status != 'Завершена'")
            open_total = cur.fetchone()["cnt"]
            cur.execute("SELECT ROUND(AVG(CURRENT_DATE - created_at)) AS avg_age FROM tasks WHERE status != 'Завершена'")
            avg_age = int(cur.fetchone()["avg_age"] or 0)
            closed_all = by_status.get("Завершена", 0)
    return jsonify({
        "open_total": open_total, "closed_all": closed_all,
        "closed_7": closed_7, "closed_30": closed_30,
        "overdue": overdue, "avg_age": avg_age,
        "by_status": by_status, "by_priority": by_priority,
        "by_project": by_project,
    })


@app.route("/api/strategy-chart")
def strategy_chart():
    import json as _json
    with get_db() as conn:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            cur.execute("""
                SELECT week_date, scores, review_text
                FROM strategic_snapshots
                ORDER BY week_date DESC
                LIMIT 12
            """)
            rows = cur.fetchall()
    rows = list(reversed(rows))  # хронологический порядок
    weeks = [str(r['week_date']) for r in rows]
    reviews = {str(r['week_date']): r['review_text'] for r in rows}
    series = {area: [] for area in STRATEGIC_AREAS}
    for r in rows:
        sc = r['scores'] if isinstance(r['scores'], dict) else _json.loads(r['scores'])
        for area in STRATEGIC_AREAS:
            val = sc.get(area)
            series[area].append(val if isinstance(val, (int, float)) else None)
    return jsonify({"weeks": weeks, "series": series, "reviews": reviews})


@app.route("/api/calendar-tasks")
def calendar_tasks():
    _PRIORITY_COLOR = {"Высокий": "#ff3b30", "Средний": "#ff9500", "Низкий": "#34c759"}
    with get_db() as conn:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            cur.execute("""
                SELECT id, title, priority, category, status, deadline, assignee, energy, progress
                FROM tasks
                WHERE status != 'Завершена' AND deadline IS NOT NULL
                ORDER BY deadline
            """)
            tasks = cur.fetchall()
    events = []
    for t in tasks:
        events.append({
            "id": t["id"],
            "title": t["title"],
            "start": str(t["deadline"]),
            "color": _PRIORITY_COLOR.get(t["priority"], "#aeaeb2"),
            "textColor": "white",
            "extendedProps": {
                "priority": t["priority"] or "",
                "category": t["category"] or "",
                "status": t["status"] or "",
                "assignee": t["assignee"] or "",
                "energy": t["energy"] or "",
                "progress": t["progress"] or "",
            }
        })
    return jsonify(events)


@app.route("/digest", methods=["POST"])
def digest_manual():
    digest_type = request.json.get("type", "daily") if request.is_json else "daily"
    try:
        build_digest(digest_type)
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


@app.route("/analyze", methods=["POST"])
def analyze():
    today = date.today().isoformat()

    with get_db() as conn:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            cur.execute("SELECT title, status, priority, deadline FROM tasks ORDER BY created_at")
            tasks = cur.fetchall()

    if not tasks:
        return jsonify({"summary": "Задач пока нет — добавьте первую!"})

    tasks_text = "\n".join([
        f"- «{t['title']}» | статус: {t['status']} | приоритет: {t['priority']} | дедлайн: {t['deadline'] or 'не указан'}"
        for t in tasks
    ])

    response = openai_client.chat.completions.create(
        model="gpt-4o",
        messages=[
            {
                "role": "system",
                "content": "Ты помощник-менеджер задач. Отвечай на русском языке, кратко и структурированно."
            },
            {
                "role": "user",
                "content": f"Сегодня {today}. Вот список задач:\n{tasks_text}\n\n"
                           f"Сделай резюме: что выполнено, что просрочено, что в работе, "
                           f"что ещё не начато. Предложи как сгруппировать задачи по смыслу."
            }
        ]
    )

    summary = response.choices[0].message.content
    return jsonify({"summary": summary})


@app.route("/subtask/add", methods=["POST"])
def subtask_add():
    task_id = request.form.get("task_id", type=int)
    text = request.form.get("text", "").strip()
    if task_id and text:
        with get_db() as conn:
            with conn.cursor() as cur:
                cur.execute("INSERT INTO subtasks (task_id, text) VALUES (%s, %s)", (task_id, text))
    return redirect(url_for("edit", task_id=task_id))


@app.route("/subtask/toggle/<int:sub_id>", methods=["POST"])
def subtask_toggle(sub_id):
    with get_db() as conn:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            cur.execute("SELECT task_id, done FROM subtasks WHERE id=%s", (sub_id,))
            row = cur.fetchone()
            if row:
                cur.execute("UPDATE subtasks SET done=%s WHERE id=%s", (not row["done"], sub_id))
                return redirect(url_for("edit", task_id=row["task_id"]))
    return redirect(url_for("index"))


@app.route("/subtask/delete/<int:sub_id>", methods=["POST"])
def subtask_delete(sub_id):
    with get_db() as conn:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            cur.execute("SELECT task_id FROM subtasks WHERE id=%s", (sub_id,))
            row = cur.fetchone()
            cur.execute("DELETE FROM subtasks WHERE id=%s", (sub_id,))
            if row:
                return redirect(url_for("edit", task_id=row["task_id"]))
    return redirect(url_for("index"))


@app.route("/dashboard")
def dashboard():
    today = date.today()
    week_ago = today - timedelta(days=7)
    month_ago = today - timedelta(days=30)

    with get_db() as conn:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            cur.execute("SELECT status, COUNT(*) AS cnt FROM tasks GROUP BY status")
            by_status = {r["status"]: r["cnt"] for r in cur.fetchall()}

            cur.execute("SELECT priority, COUNT(*) AS cnt FROM tasks WHERE status != 'Завершена' GROUP BY priority")
            by_priority = {r["priority"]: r["cnt"] for r in cur.fetchall()}

            cur.execute("SELECT COUNT(*) AS cnt FROM tasks WHERE status='Завершена' AND closed_at >= %s", (week_ago,))
            closed_7 = cur.fetchone()["cnt"]

            cur.execute("SELECT COUNT(*) AS cnt FROM tasks WHERE status='Завершена' AND closed_at >= %s", (month_ago,))
            closed_30 = cur.fetchone()["cnt"]

            cur.execute("SELECT COUNT(*) AS cnt FROM tasks WHERE deadline < %s AND status != 'Завершена'", (today,))
            overdue = cur.fetchone()["cnt"]

            cur.execute("""
                SELECT project, COUNT(*) AS cnt
                FROM tasks WHERE status != 'Завершена' AND project IS NOT NULL
                GROUP BY project ORDER BY cnt DESC LIMIT 8
            """)
            by_project = cur.fetchall()

            cur.execute("SELECT COUNT(*) AS cnt FROM tasks WHERE status != 'Завершена'")
            open_total = cur.fetchone()["cnt"]

            cur.execute("SELECT ROUND(AVG(CURRENT_DATE - created_at)) AS avg_age FROM tasks WHERE status != 'Завершена'")
            avg_age = cur.fetchone()["avg_age"] or 0

    return render_template("dashboard.html",
        by_status=by_status,
        by_priority=by_priority,
        closed_7=closed_7,
        closed_30=closed_30,
        overdue=overdue,
        by_project=by_project,
        open_total=open_total,
        avg_age=int(avg_age),
        today=today,
    )


@app.route("/static/sw.js")
def service_worker():
    return send_from_directory("static", "sw.js", mimetype="application/javascript")


@app.route("/health")
def health():
    result = {}
    result["DATABASE_URL_set"] = bool(os.environ.get("DATABASE_URL"))
    result["OPENAI_API_KEY_set"] = bool(os.environ.get("OPENAI_API_KEY"))
    result["TELEGRAM_configured"] = bool(TELEGRAM_TOKEN and TELEGRAM_CHAT_ID)
    result["DATABASE_URL_prefix"] = (os.environ.get("DATABASE_URL") or "")[:30]
    try:
        with get_db() as conn:
            with conn.cursor() as cur:
                cur.execute("SELECT 1")
        result["db"] = "OK"
    except Exception as e:
        result["db"] = str(e)
    return jsonify(result)


@app.errorhandler(500)
def server_error(e):
    import traceback
    return f"<pre>{traceback.format_exc()}</pre>", 500


try:
    init_db()
except Exception as e:
    print(f"init_db error: {e}")

if __name__ == "__main__":
    app.run()
